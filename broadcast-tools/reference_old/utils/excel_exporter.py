import pandas as pd
import logging
from typing import List, Dict, Any, Optional
import os
from openpyxl import load_workbook
import errno
from pathlib import Path

logger = logging.getLogger(__name__)

class ExcelExporter:
    @staticmethod
    def check_file_access(filepath: str) -> tuple[bool, str]:
        """Check if file can be accessed/written to"""
        try:
            # Try to open the file in append mode to check write access
            if os.path.exists(filepath):
                with open(filepath, 'a+b') as f:
                    f.seek(0)  # Try to read
                    f.seek(0, 2)  # Try to write
            return True, ""
        except IOError as e:
            if e.errno == errno.EACCES:
                return False, "File is open in another program or requires admin access"
            elif e.errno == errno.EAGAIN:
                return False, "File is locked by another process"
            elif e.errno == errno.ENOENT:
                # File doesn't exist - this is fine for new files
                return True, ""
            else:
                return False, f"Unable to access file: {str(e)}"
        except Exception as e:
            return False, str(e)

    @staticmethod
    def export_to_new_excel(drawing_data: List[Dict[str, Any]], filepath: str):
        """Export drawing data to a new Excel file"""
        try:
            df = pd.DataFrame(drawing_data)
            df.to_excel(filepath, sheet_name='Imported Data', index=False)
            logger.info(f"Data exported to new file: {filepath}")
        except Exception as e:
            logger.exception("Error creating new Excel file:")
            raise

    @staticmethod
    def merge_to_excel(drawing_data: List[Dict[str, Any]], filepath: str, 
                      sheet_name: Optional[str] = None):
        """Merge drawing data with existing Excel file"""
        try:
            # Check file access first
            can_access, error_msg = ExcelExporter.check_file_access(filepath)
            if not can_access:
                raise PermissionError(f"Cannot access Excel file: {error_msg}")

            new_df = pd.DataFrame(drawing_data)
            
            try:
                # Try to load workbook
                book = load_workbook(filepath)
            except PermissionError:
                raise PermissionError("Excel file is currently open. Please close it and try again.")
            except Exception as e:
                raise Exception(f"Error reading Excel file: {str(e)}")

            # Rest of the merge logic...
            if sheet_name is None:
                sheets = book.sheetnames
                if 'Imported Data' in sheets:
                    sheet_name = 'Imported Data'
                else:
                    sheet_name = sheets[0]

            existing_df = pd.read_excel(filepath, sheet_name=sheet_name)
            merged_df = pd.concat([existing_df, new_df], ignore_index=True)
            merged_df = merged_df.drop_duplicates(subset=['DWG'], keep='last')

            try:
                with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', 
                                  if_sheet_exists='replace') as writer:
                    merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except PermissionError:
                raise PermissionError("Cannot write to Excel file. Please ensure it is not open in Excel.")
            
            logger.info(f"Data merged into {filepath}, sheet: {sheet_name}")
            
        except Exception as e:
            logger.exception("Error merging with Excel file:")
            raise

    @staticmethod
    def get_available_sheets(filepath: str) -> List[str]:
        """Get list of available worksheets in Excel file"""
        try:
            book = load_workbook(filepath, read_only=True)
            return book.sheetnames
        except Exception as e:
            logger.exception("Error reading Excel sheets:")
            raise
